"""DRE calculation and Excel export."""
from __future__ import annotations
from io import BytesIO

import pandas as pd

from config import DRE_ROWS, MONTHS_PT


def calculate_dre(transactions: list[dict]) -> list[dict]:
    """
    Given a list of transaction dicts, return DRE rows ready for display.
    Each row: {"label", "amount", "type", "level"}
    """
    df = pd.DataFrame(transactions) if transactions else pd.DataFrame(
        columns=["category", "amount"]
    )

    def sum_cats(categories: list[str]) -> float:
        if df.empty:
            return 0.0
        mask = df["category"].isin(categories)
        total = df.loc[mask, "amount"].sum()
        return float(total)

    # First pass: compute raw values per row id
    values: dict[str, float] = {}
    for row in DRE_ROWS:
        if row["type"] == "section":
            continue
        if "categories" in row:
            raw = sum_cats(row["categories"])
            # Income rows keep sign; cost rows show absolute value
            values[row["id"]] = raw if row["type"] == "income" else abs(raw)
        elif "formula" in row:
            values[row["id"]] = row["formula"](values)

    # Second pass: build display rows
    result = []
    for row in DRE_ROWS:
        amount = values.get(row["id"])
        result.append({
            "label":  row["label"],
            "amount": amount,
            "type":   row["type"],
            "level":  row["level"],
        })
    return result


def export_excel(dre_rows: list[dict], month: int, year: int) -> bytes:
    """Return Excel file bytes with a formatted DRE."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback: plain CSV-style Excel via pandas
        df = _to_df(dre_rows)
        buf = BytesIO()
        df.to_excel(buf, index=False)
        return buf.getvalue()

    wb = openpyxl.Workbook()
    ws = wb.active
    month_name = MONTHS_PT.get(month, str(month))
    ws.title = f"DRE {month_name[:3]} {year}"

    # Styles
    PURPLE = "6B21A8"
    LIGHT_PURPLE = "F3E8FF"
    GREEN = "166534"
    RED_BG = "FEE2E2"
    GRAY = "F3F4F6"

    def cell_style(ws, row, col, value, bold=False, bg=None, color=None,
                   align="left", number_fmt=None, indent=0):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=color or "000000", size=11)
        c.alignment = Alignment(horizontal=align, indent=indent)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        if number_fmt:
            c.number_format = number_fmt
        return c

    # Header
    ws.merge_cells("A1:B1")
    ws["A1"] = "Florence Intimates — DRE"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=PURPLE)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:B2")
    ws["A2"] = f"{month_name} / {year}"
    ws["A2"].font = Font(bold=False, size=11, color="6B21A8")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[3].height = 6

    header_row = 4
    ws.cell(row=header_row, column=1, value="DESCRIÇÃO").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=header_row, column=2, value="VALOR (R$)").font = Font(bold=True, color="FFFFFF")
    for col in (1, 2):
        c = ws.cell(row=header_row, column=col)
        c.fill = PatternFill("solid", fgColor=PURPLE)
        c.alignment = Alignment(horizontal="center")

    r = header_row + 1
    for entry in dre_rows:
        t     = entry["type"]
        label = entry["label"]
        amt   = entry["amount"]
        level = entry["level"]

        if t == "section":
            ws.merge_cells(f"A{r}:B{r}")
            cell_style(ws, r, 1, label, bold=True, bg=LIGHT_PURPLE, color=PURPLE)
            ws.row_dimensions[r].height = 20
        elif t == "income":
            cell_style(ws, r, 1, label, indent=level)
            cell_style(ws, r, 2, amt, align="right",
                       number_fmt='R$ #,##0.00', color=GREEN)
        elif t == "cost":
            cell_style(ws, r, 1, label, indent=level, bg=GRAY if level > 0 else None)
            display = -abs(amt) if amt != 0 else 0
            cell_style(ws, r, 2, display, align="right",
                       number_fmt='R$ #,##0.00', bg=GRAY if level > 0 else None)
        elif t == "subtotal":
            is_result = label.startswith("=")
            bg_color = "DCFCE7" if (amt or 0) >= 0 else "FEE2E2"
            cell_style(ws, r, 1, label, bold=True)
            cell_style(ws, r, 2, amt, bold=True, align="right",
                       number_fmt='R$ #,##0.00', bg=bg_color)
        elif t == "result":
            bg = "DCFCE7" if (amt or 0) >= 0 else "FEE2E2"
            cell_style(ws, r, 1, label, bold=True, bg=bg)
            cell_style(ws, r, 2, amt, bold=True, align="right",
                       number_fmt='R$ #,##0.00', bg=bg)

        r += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_df(dre_rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame([
        {"Descrição": r["label"], "Valor (R$)": r.get("amount")}
        for r in dre_rows if r["type"] != "section"
    ])
